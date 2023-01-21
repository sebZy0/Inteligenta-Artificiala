import openpyxl as px
from PyQt6.QtCore import Qt, QAbstractTableModel, QVariant, QModelIndex
from PyQt6.QtGui import QIcon
from PyQt6 import QtCore, QtGui, QtWidgets, uic
import sys
import os
from PyQt6.QtWidgets import QWidget, QApplication, QMainWindow, QLabel, QLineEdit, QPushButton, QHBoxLayout, QVBoxLayout, QTableWidget, QTableWidgetItem, QComboBox

# currency_formats = ['#,##0.00[$-409]', '#,##0.00[$-404]',
#                     '#,##0.00[$-x-symbol]', '#,##0.00[$-en-US]', '#,##0.00 "RON"', '#,##0.00 "EUR"', '#,##0.00 "GBP"', '#,##0.00 "USD"',
#                     '"€" #,##.00', '"RON" #,##.00', '"£" #,##.00', '"$" #,##.00', 'RON', 'EUR', 'GBP', 'USD']


class BudgetApp(QWidget):
    def __init__(self):
        super().__init__()
        uic.loadUi('landing.ui', self)
        self.w = None
        self.setWindowTitle("Budget manager")
        self.setWindowIcon(QIcon("icon2.ico"))
        self.resize(600, 400)

        importButton = self.findChild(QPushButton, 'importButton')
        importButton.clicked.connect(self.clickedImportButton)

        exportButton = self.findChild(QPushButton, 'exportButton')
        exportButton.clicked.connect(self.clickedExportButton)

        addRowButton = self.findChild(QPushButton, 'addRowButton')
        addRowButton.clicked.connect(self.addNewRow)

        comboBox = self.findChild(QComboBox, "comboBox")
        comboBox.currentTextChanged.connect(self.comboChanged)

        incomeTableButton = self.findChild(QPushButton, 'incomeButton')
        incomeTableButton.clicked.connect(self.createIncomeTable)

        expensesTableButton = self.findChild(QPushButton, 'expensesButton')
        expensesTableButton.clicked.connect(self.createExpensesTable)

    def clickedImportButton(self):
        global file_path
        file_dialog = QtWidgets.QFileDialog()
        file_dialog.setNameFilter("Excel files (*.xlsx *.xls)")
        file_path, _ = file_dialog.getOpenFileName()

        if file_path.endswith(('.xlsx', '.xls')):
            self.loadTable()
        else:
            print("Not an Excel file.")
        income = 0
        expenses = 0

    def loadTable(self):
        wb = px.load_workbook(file_path)
        ws = wb['Sheet1']
        global table
        table = self.findChild(QTableWidget, "tableWidget")
        table.setRowCount(ws.max_row)
        table.setColumnCount(5)
        global total
        total = 0
        self.income = 0
        self.expenses = 0
        operator = 1
        for i in range(1, ws.max_row+1):
            for j in range(1, 6):
                cell = ws.cell(row=i, column=j)
                item = QTableWidgetItem(str(cell.value))
                # if j == 6:
                #     item.setTextAlignment(
                #         Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                #     item.setData(Qt.ItemDataRole.UserRole,
                #                  currency_format)
                if cell.value:
                    # currency_format = ws['E2'].number_format
                    # print(currency_format)
                    item = QTableWidgetItem(str(cell.value))
                    if i >= 2:
                        if j == 2:
                            if (str(cell.value) == "Income"):
                                operator = 1
                            else:
                                operator = 0
                        if (j == 5):
                            if operator == 1:
                                self.income = self.income + int(cell.value)
                                total = total + int(cell.value)
                            else:
                                self.expenses = self.expenses + int(cell.value)
                                total = total - int(cell.value)
                    table.setItem(i-1, j-1, item)
        print(total)
        totalSum = self.findChild(QLabel, "totalSum")
        incomeLabel = self.findChild(QLabel, "incomeLabel")
        expensesLabel = self.findChild(QLabel, "expensesLabel")
        currency = self.comboBox.currentText()
        totalSum.setText("Total: " + str(total) + " " + currency)
        incomeLabel.setText(str(self.income) + currency)
        expensesLabel.setText(str(self.expenses) + currency)
        delete_unpopulated_rows(table)
        table.show()

    def toggle_window(self, window):
        if window.isVisible():
            window.hide()
        else:
            window.show()

    def comboChanged(self):
        totalSum = self.findChild(QLabel, "totalSum")
        item = self.comboBox.currentText()
        totalSum.setText("Total: " + str(total) + " " + item)

    def createIncomeTable(self):
        if self.w is None:
            self.w = IncomeWindow()
            self.w.show()
        else:
            self.w.close()
            self.w = None

    def createExpensesTable(self):
        if self.w is None:
            self.w = ExpensesWindow()
            self.w.show()
        else:
            self.w.close()
            self.w = None

    def clickedExportButton(self):
        table = self.findChild(QTableWidget, "tableWidget")
        file_dialog = QtWidgets.QFileDialog()
        file_dialog.getSaveFileName(directory='table.xlsx')
        file_dialog.setNameFilter("Excel files (*.xlsx *.xls)")
        wb = px.Workbook()
        ws = wb.active
        ws.title = 'Sheet1'
        for row in range(table.rowCount()):
            for column in range(table.columnCount()):
                item = table.item(row, column)
                if item:
                    ws.cell(row=row+1, column=column+1).value = item.text()
        wb.save(file_path)
        self.loadTable()

    def addNewRow(self):
        table = self.findChild(QTableWidget, "tableWidget")
        table.insertRow(table.rowCount())

    def getTable(self):
        table = self.findChild(QTableWidget, "tableWidget")
        return table


class IncomeWindow(QWidget):
    def __init__(self):
        super().__init__()
        uic.loadUi('income.ui', self)
        self.setWindowTitle("Budget manager - Income")
        self.setWindowIcon(QIcon("icon2.ico"))

        self.createIncomeTable()

    def createIncomeTable(self):
        global table
        income_table = self.findChild(QTableWidget, "incomeTable")
        income_table.setRowCount(table.rowCount())
        income_table.setColumnCount(table.columnCount())
        row = 0
        for i in range(0, table.rowCount()):
            if table.item(i, 1).text() == "Income":
                row = row + 1
                for j in range(table.columnCount()):
                    item = QTableWidgetItem(table.item(i, j).text())
                    income_table.setItem(row, j, item)

        income_table.show()


class ExpensesWindow(QWidget):
    def __init__(self):
        super().__init__()
        uic.loadUi('expenses.ui', self)
        self.setWindowTitle("Budget manager - Expenses")
        self.setWindowIcon(QIcon("icon2.ico"))
        self.createExpensesTable()

    def createExpensesTable(self):
        global table
        expenses_table = self.findChild(QTableWidget, "expensesTable")
        expenses_table.setRowCount(table.rowCount())
        expenses_table.setColumnCount(table.columnCount())
        row = 0
        for i in range(0, table.rowCount()):
            if table.item(i, 1).text() == "Expense":
                row = row + 1
                for j in range(table.columnCount()):
                    item = QTableWidgetItem(table.item(i, j).text())
                    expenses_table.setItem(row, j, item)
        expenses_table.show()


def delete_unpopulated_rows(table):
    for i in range(table.rowCount()):
        row_empty = True
        for j in range(table.columnCount()):
            item = table.item(i, j)
            if item and item.text():
                row_empty = False
                break
        if row_empty:
            table.removeRow(i)
            i -= 1


app = QApplication(sys.argv)
window = BudgetApp()
window.show()
app.exec()
